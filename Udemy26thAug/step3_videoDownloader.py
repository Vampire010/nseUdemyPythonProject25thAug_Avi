import os
import re
import subprocess
import requests
from openpyxl import load_workbook

def safe_name(name: str) -> str:
    s = re.sub(r'[<>:"/\\|?*\n\r\t]', "_", str(name))
    s = s.strip()
    s = re.sub(r"_+", "_", s)
    s = s.rstrip(" .")
    reserved = {
        "CON","PRN","AUX","NUL",
        "COM1","COM2","COM3","COM4","COM5","COM6","COM7","COM8","COM9",
        "LPT1","LPT2","LPT3","LPT4","LPT5","LPT6","LPT7","LPT8","LPT9"
    }
    if s.upper() in reserved or s == "":
        s = f"_{s}_" if s else "Untitled"
    if len(s) > 120:
        s = s[:120]
    return s

def get_video_folder(base_dir, course_name, section_idx, section_name, lecture_idx, lecture_name):
    course_folder = safe_name(course_name)
    section_folder = f"{section_idx:02d}_{safe_name(section_name)}"
    lecture_folder = f"{lecture_idx:02d}_{safe_name(lecture_name)}"
    full_path = os.path.join(base_dir, course_folder, section_folder, lecture_folder)
    os.makedirs(full_path, exist_ok=True)
    return full_path

def fetch_video_playlist(asset_id):
    url = f"https://www.udemy.com/assets/{asset_id}/files/2021-08-07_08-10-23-f0420cdf436995091bef63c49990323a/2/aa00d84944423034059058c0302b6aa3a4bb.m3u8?token=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJwYXRoIjoiMjAyMS0wOC0wN18wOC0xMC0yMy1mMDQyMGNkZjQzNjk5NTA5MWJlZjYzYzQ5OTkwMzIzYS8yLyIsImV4cCI6MTc1NjU0OTQxNH0.JC8-65AzLFs9AlJyGdn-x3PM4jps0JbpikJSp8ETHkE&provider=cloudfront&v=1"
    headers = {
        "accept": "*/*",
        "accept-language": "en-GB,en;q=0.9",
        "priority": "u=1, i",
        "sec-ch-ua": "\"Not;A=Brand\";v=\"99\", \"Google Chrome\";v=\"139\", \"Chromium\";v=\"139\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\"",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "cookie": "__udmy_2_v57r=53e18217f3d94578b7f5805607a3bd40; __cf_bm=0XYa1VwzkuSE8FCWmKnpOAb74aLA0POLrHWu_l3zxvQ-1756451165-1.0.1.1-t5yf5sYkCA3gbVGcUxSmowwg0TqE5E.5NVvHHsTN.qZlBcus929iz871GEP_BTxWOUsqxvAwqq8kjqZErRpJBBYha1O0HYP9ZNGGLQ09zGU; ..."
    }
    response = requests.get(url, headers=headers)
    if response.ok:
        return response.content.decode("utf-8")
    else:
        response.raise_for_status()

def extract_highest_bandwidth_url(playlist):
    lines = playlist.splitlines()
    highest_bandwidth = 0
    highest_url = ""
    for i, line in enumerate(lines):
        if line.startswith("#EXT-X-STREAM-INF"):
            bandwidth_match = re.search(r'BANDWIDTH=(\d+)', line)
            if bandwidth_match:
                bandwidth = int(bandwidth_match.group(1))
                if bandwidth > highest_bandwidth and i + 1 < len(lines):
                    highest_bandwidth = bandwidth
                    highest_url = lines[i + 1].strip()
    return highest_url

def download_video(m3u8_url, output_file):
    ffmpeg_headers = (
        "Accept: */*\r\n"
        "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64)\r\n"
        "Referer: https://www.udemy.com/\r\n"
        "Cookie: __udmy_2_v57r=53e18217f3d94578b7f5805607a3bd40; __cf_bm=0XYa1VwzkuSE8FCWmKnpOAb74aLA0POLrHWu_l3zxvQ-1756451165-1.0.1.1-t5yf5sYkCA3gbVGcUxSmowwg0TqE5E.5NVvHHsTN.qZlBcus929iz871GEP_BTxWOUsqxvAwqq8kjqZErRpJBBYha1O0HYP9ZNGGLQ09zGU; ...\r\n"
    )
    command = ["ffmpeg", "-headers", ffmpeg_headers, "-i", m3u8_url, "-c", "copy", output_file]
    try:
        subprocess.run(command, check=True)
        print(f"Downloaded: {output_file}")
    except subprocess.CalledProcessError as e:
        print(f"Failed to download: {output_file}. Error: {e}")
        raise

def process_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_mapping = {str(col).strip().lower(): idx for idx, col in enumerate(header_row) if col is not None}
    required_columns = {
        "course_name": None,
        "chapter_title": None,
        "item_title": None,
        "asset_id": None
    }
    for key in required_columns:
        if key in header_mapping:
            required_columns[key] = header_mapping[key]
        else:
            print(f"Required column '{key}' not found in the header.")
            return

    chapter_indices = {}
    chapter_counter = 0
    lecture_counter = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        course_name = row[required_columns["course_name"]]
        chapter_title = row[required_columns["chapter_title"]]
        lecture_title = row[required_columns["item_title"]]
        asset_id = row[required_columns["asset_id"]]

        if asset_id is None or str(asset_id).strip() == "":
            continue

        if chapter_title not in chapter_indices:
            chapter_counter += 1
            chapter_indices[chapter_title] = chapter_counter
        chapter_number = chapter_indices[chapter_title]

        folder_path = get_video_folder("./udemyDownloads/videoDownloaded", course_name, chapter_number, chapter_title, lecture_counter, lecture_title)
        print(f"Processing: {course_name} /{chapter_number:02d}_{chapter_title} / {lecture_counter:02d}_{lecture_title} with asset ID: {asset_id}")

        try:
            playlist = fetch_video_playlist(asset_id)
            highest_bandwidth_url = extract_highest_bandwidth_url(playlist)
            if highest_bandwidth_url:
                output_file = os.path.join(folder_path, f"{lecture_counter:02d}_{safe_name(lecture_title)}.mp4")
                print("Downloading from URL with highest bandwidth:")
                print(highest_bandwidth_url)
                for attempt in range(1, 4):
                    try:
                        download_video(highest_bandwidth_url, output_file)
                        break
                    except Exception as e:
                        print(f"Attempt {attempt} failed: {e}")
                        if attempt == 3:
                            print(f"Failed to download after 3 attempts: {output_file}")
            else:
                print("No valid stream URL found for asset ID:", asset_id)
        except Exception as e:
            print(f"Error processing asset_id {asset_id}: {e}")

        lecture_counter += 1

if __name__ == "__main__":
    excel_file_path = r"./udemyDownloads/testDataCourse.xlsx"
    process_excel(excel_file_path)


'''

video 1 = https://www.udemy.com/api-2.0/users/me/subscribed-courses



Video2 =



'''