import requests
import time
import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from getCookiesFromNSEIndia import NSECookieManager  # Import NSECookieManager

# Constants
EXCEL_FILE = r"./exportedData/OptionChainData.xlsx"
CHECK_INTERVAL = 30  # seconds

class NSEDataFetcher:
    def __init__(self):
        # Step 1: Fetch cookies
        try:
            cm = NSECookieManager()
            self.cookies = cm.fetch_and_save_cookies()
        except Exception as e:
            print(f"[ERROR] Failed to fetch cookies: {e}")
            self.cookies = {}

        # Step 2: Load saved cookies
        cookie_file = "nseIndiaCookies_name_value.json"
        try:
            if os.path.exists(cookie_file):
                with open(cookie_file, "r") as f:
                    auth_data = json.load(f)
            else:
                print(f"[WARNING] Cookie file not found: {cookie_file}")
                auth_data = {}
        except Exception as e:
            print(f"[ERROR] Failed to load cookie JSON: {e}")
            auth_data = {}

        # Safe cookie get
        def safe_get(key):
            return auth_data.get(key, "")

        cookie_str = ";".join([f"{k}={v}" for k, v in auth_data.items()])
        self.url = "https://www.nseindia.com/api/live-analysis-oi-spurts-underlyings"
        self.headers = self._make_headers(cookie_str)

        # Step 3: Output location
        export_dir = r"./exportedData"
        os.makedirs(export_dir, exist_ok=True)
        self.output_file = os.path.join(export_dir, "OptionChainData.xlsx")

    def _make_headers(self, cookie_str):
        return {
            'accept': '*/*',
            'accept-language': 'en-GB,en-IN;q=0.9,en-US;q=0.8,en;q=0.7',
            'priority': 'u=1, i',
            'referer': 'https://www.nseindia.com/market-data/oi-spurts',
            'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
            'Cookie': cookie_str
        }

    # Fetch data from API
    def fetch_data(self):
        response = requests.get(self.url, headers=self.headers)
        response.raise_for_status()
        return response.json()

# Initialize Excel file
def initialize_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(EXCEL_FILE)

# Write data to Excel with dynamic attributes as column headers
def write_to_excel(data, sheet_name):
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        initialize_excel()
        wb = load_workbook(EXCEL_FILE)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Extract column headers dynamically from the first data item
    if data:
        headers = list(data[0].keys())
        if ws.max_row == 1:  # Write headers only if the sheet is empty
            ws.append(headers)

        # Append data rows
        for item in data:
            ws.append([item.get(header, "") for header in headers])

    wb.save(EXCEL_FILE)

# Highlight changes in Excel
def highlight_changes(new_data, old_data, sheet_name):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Get column headers
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        symbol = row[headers.index("symbol")].value
        if symbol in new_data:
            for col_idx, header in enumerate(headers):
                new_value = new_data[symbol].get(header, None)
                old_value = old_data.get(symbol, {}).get(header, None)
                if new_value != old_value:
                    row[col_idx].fill = red_fill
                else:
                    row[col_idx].fill = green_fill

    wb.save(EXCEL_FILE)

# Extract all fields dynamically from the response
def extract_data(response):
    extracted_data = {}
    for item in response.get("data", []):
        symbol = item.get("symbol", "")
        extracted_data[symbol] = item
    return extracted_data

# Main loop
def main():
    print("Starting Option Chain Monitor...")
    fetcher = NSEDataFetcher()
    previous_data = {}

    while True:
        try:
            print("Fetching data...")
            response = fetcher.fetch_data()

            # Extract and structure data
            current_data = extract_data(response)

            # Write to Excel
            timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
            write_to_excel(list(current_data.values()), f"Sheet_{timestamp}")

            # Highlight changes
            if previous_data:
                highlight_changes(current_data, previous_data, f"Sheet_{timestamp}")

            previous_data = current_data
            print(f"Data written to Excel at {timestamp}. Waiting for {CHECK_INTERVAL} seconds...")
        except Exception as e:
            print(f"Error: {e}")

        time.sleep(CHECK_INTERVAL)

if __name__ == "__main__":
    main()