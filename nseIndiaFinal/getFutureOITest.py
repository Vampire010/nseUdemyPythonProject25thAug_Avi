import requests
import time
import os
import json
from openpyxl import Workbook, load_workbook
from getCookiesFromNSEIndia import NSECookieManager  # Import NSECookieManager

# Constants
EXCEL_FILE = r"./exportedData/OptionChainDataTest.xlsx"
CHECK_INTERVAL = 30  # seconds

class NSEDataFetcher:
    def __init__(self):
        # Step 1: Fetch cookies
        try:
            cm = NSECookieManager()
            self.cookies = cm.fetch_and_save_cookies()
        except Exception as e:
            print(f"[ERROR] Failed to fetch cookies: {e}")
            self.cookies = {}

        # Step 2: Load saved cookies
        cookie_file = "nseIndiaCookies_name_value.json"
        try:
            if os.path.exists(cookie_file):
                with open(cookie_file, "r") as f:
                    auth_data = json.load(f)
            else:
                print(f"[WARNING] Cookie file not found: {cookie_file}")
                auth_data = {}
        except Exception as e:
            print(f"[ERROR] Failed to load cookie JSON: {e}")
            auth_data = {}

        cookie_str = ";".join([f"{k}={v}" for k, v in auth_data.items()])
        self.url = "https://www.nseindia.com/api/live-analysis-oi-spurts-underlyings"
        self.headers = self._make_headers(cookie_str)

        export_dir = r"./exportedData"
        os.makedirs(export_dir, exist_ok=True)
        self.output_file = os.path.join(export_dir, "OptionChainData.xlsx")

    def _make_headers(self, cookie_str):
        return {
            'accept': '*/*',
            'accept-language': 'en-GB,en-IN;q=0.9,en-US;q=0.8,en;q=0.7',
            'priority': 'u=1, i',
            'referer': 'https://www.nseindia.com/market-data/oi-spurts',
            'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
            'Cookie': cookie_str
        }

    def fetch_data(self):
        response = requests.get(self.url, headers=self.headers)
        response.raise_for_status()
        return response.json()

def initialize_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(EXCEL_FILE)

def get_prev_sheet_latestOI(sheet_name):
    wb = load_workbook(EXCEL_FILE)
    sheetnames = wb.sheetnames
    idx = sheetnames.index(sheet_name) if sheet_name in sheetnames else len(sheetnames)
    if idx == 0:
        return None  # No previous sheet
    prev_sheet = wb[sheetnames[idx - 1]]
    prev_data = {}
    for row in prev_sheet.iter_rows(min_row=2, values_only=True):
        symbol = row[0]
        latestOI = row[1]
        prev_data[symbol] = latestOI
    return prev_data

def write_to_excel_with_oi(data, sheet_name, prev_sheet_data=None):
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        initialize_excel()
        wb = load_workbook(EXCEL_FILE)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Write headers if the sheet is empty
    if ws.max_row == 1:
        ws.append(["symbol", "latestOI", "prevOI", "diffInLatestOI"])

    # Write data rows
    for symbol, item in data.items():
        latestOI = item.get("latestOI", "")
        prevOI = item.get("prevOI", "")
        diffInLatestOI = ""
        if prev_sheet_data and symbol in prev_sheet_data:
            prev_latestOI = prev_sheet_data[symbol]
            if latestOI != "" and prev_latestOI != "" and prev_latestOI is not None and latestOI is not None:
                try:
                    diffInLatestOI = float(prev_latestOI) - float(latestOI)
                except Exception:
                    diffInLatestOI = ""
        ws.append([
            symbol,
            latestOI,
            prevOI,
            diffInLatestOI
        ])

    wb.save(EXCEL_FILE)

def extract_data(response):
    extracted_data = {}
    for item in response.get("data", []):
        symbol = item.get("symbol", "")
        extracted_data[symbol] = {
            "symbol": symbol,
            "latestOI": item.get("latestOI"),
            "prevOI": item.get("prevOI")
        }
    return extracted_data

def main():
    print("Starting Option Chain Monitor...")
    fetcher = NSEDataFetcher()

    while True:
        try:
            print("Fetching data...")
            response = fetcher.fetch_data()
            current_data = extract_data(response)

            timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
            sheet_name = f"Sheet_{timestamp}"
            prev_sheet_data = get_prev_sheet_latestOI(sheet_name)
            write_to_excel_with_oi(current_data, sheet_name, prev_sheet_data)

            print(f"Data written to Excel at {timestamp}. Waiting for {CHECK_INTERVAL} seconds...")
        except Exception as e:
            print(f"Error: {e}")

        time.sleep(CHECK_INTERVAL)

if __name__ == "__main__":
    main()