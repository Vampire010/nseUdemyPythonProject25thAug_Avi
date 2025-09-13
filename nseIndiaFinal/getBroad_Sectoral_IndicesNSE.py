import requests
import pandas as pd
import time
import json
import os
from openpyxl.styles import PatternFill
from getCookiesFromNSEIndia import NSECookieManager
from concurrent.futures import ThreadPoolExecutor


class NseTestDataExporter:

    def __init__(self):
        # --- Step 1: Get cookies safely ---
        try:
            cm = NSECookieManager()
            self.cookies = cm.fetch_and_save_cookies()
        except Exception as e:
            print(f"[ERROR] Failed to fetch cookies: {e}")
            self.cookies = {}

        self.marketIndices = [
            "Broad Market Indices",
            "Sectoral Indices"
        ]

        # --- Step 2: Load saved cookies from file ---
        cookie_file = "nseIndiaCookies_name_value.json"
        if os.path.exists(cookie_file):
            try:
                with open(cookie_file, "r") as f:
                    auth_data = json.load(f)
            except Exception as e:
                print(f"[ERROR] Failed to load cookie JSON: {e}")
                auth_data = {}
        else:
            print(f"[WARNING] Cookie file not found: {cookie_file}")
            auth_data = {}

        # Extract cookies safely
        def safe_get(key): return auth_data.get(key, "")

        _ga = safe_get("_ga")
        abck = safe_get("_abck")
        AKA_A2 = safe_get("AKA_A2")
        nsit = safe_get("nsit")
        nseappid = safe_get("nseappid")
        bm_mi = safe_get("bm_mi")
        bm_sz = safe_get("bm_sz")
        ak_bmsc = safe_get("ak_bmsc")
        _ga_87M7PJ3R97 = safe_get("_ga_87M7PJ3R97")
        bm_sv = safe_get("bm_sv")
        RT = safe_get("RT")

        cookie_str = f'_ga={_ga};abck={abck};AKA_A2={AKA_A2};nsit={nsit};nseappid={nseappid};bm_mi={bm_mi};bm_sz={bm_sz};ak_bmsc={ak_bmsc};_ga_87M7PJ3R97={_ga_87M7PJ3R97};bm_sv={bm_sv}'

        self.broad_indices_url_template = "https://www.nseindia.com/api/heatmap-index?type={market_index}"
        self.gainers_url_template = "https://www.nseindia.com/api/heatmap-symbols?type={market_index}&indices={indices}"

        self.headers_broad_indices = self._make_headers(cookie_str)
        self.headers_gainers_url = self._make_headers(cookie_str)

        self.payload = {"gainers"}
        self.session = requests.Session()
        self.session.headers.update(self.headers_broad_indices)

    def _make_headers(self, cookie_str):
        return {
            'accept': '*/*',
            'accept-language': 'en-GB,en-IN;q=0.9,en-US;q=0.8,en;q=0.7',
            'priority': 'u=1, i',
            'referer': 'https://www.nseindia.com/market-data/live-market-indices/heatmap',
            'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
            'Cookie': cookie_str
        }

    def fetch_broad_market_indices(self, market_index):
        url = self.broad_indices_url_template.format(market_index=market_index.replace(" ", "%20"))
        try:
            response = self.session.get(url, timeout=10)
            print(f"Fetching broad market indices for: {market_index} - Status: {response.status_code}")
            response.raise_for_status()
            data = response.json()
        except requests.exceptions.RequestException as e:
            print(f"[ERROR] Request failed for {market_index}: {e}")
            return []
        except json.JSONDecodeError as e:
            print(f"[ERROR] JSON decode failed for {market_index}: {e}")
            return []

        indices_list = []
        if isinstance(data, dict):
            for value in data.values():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    indices_list = value
                    break
        elif isinstance(data, list):
            indices_list = data

        return indices_list

    def fetch_gainers_for_index(self, market_index, index_info):
        index_name = list(index_info.values())[0]
        url = self.gainers_url_template.format(
            market_index=market_index.replace(" ", "%20"),
            indices=index_name.replace(" ", "%20")
        )
        try:
            self.session.headers.update(self.headers_gainers_url)
            response = self.session.get(url, timeout=10)
            print(f"Requesting gainers for: {index_name} ({market_index}) - Status: {response.status_code}")
            response.raise_for_status()
            gainers_data = response.json()
            if isinstance(gainers_data, list):
                return [{**index_info, **row, "MarketIndex": market_index} for row in gainers_data]
        except requests.exceptions.RequestException as e:
            print(f"[ERROR] Request failed for {index_name}: {e}")
        except json.JSONDecodeError as e:
            print(f"[ERROR] JSON decode failed for {index_name}: {e}")
        return []

    def fetch_gainers_for_indices(self, market_index, indices_list):
        all_data = []
        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(self.fetch_gainers_for_index, market_index, index_info) for index_info in indices_list]
            for future in futures:
                result = future.result()
                if result:
                    all_data.extend(result)
        return all_data

    def export_to_excel(self, broad_indices_dict, gainers_dict, filename="nse_Broad_SectoralIndices_combined_data.xlsx"):
        try:
            export_dir = r"./exportedData"
            os.makedirs(export_dir, exist_ok=True)
            file_path = os.path.join(export_dir, filename)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write all data first
                for market_index in self.marketIndices:
                    broad_df = pd.DataFrame(broad_indices_dict.get(market_index, []))
                    broad_df.to_excel(writer, sheet_name=f"{market_index}_Broad"[:31], index=False)

                    gainers_df = pd.DataFrame(gainers_dict.get(market_index, []))
                    gainers_df.to_excel(writer, sheet_name=f"{market_index}_Gainers"[:31], index=False)

                # Now apply formatting to ALL sheets
                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    # Get header row
                    header_cells = [cell.value for cell in sheet[1]]
                    if not header_cells:
                        continue

                    # Find pChange column
                    pchange_col_idx = None
                    for idx, h in enumerate(header_cells):
                        if h and str(h).strip().lower() == "pchange":
                            pchange_col_idx = idx + 1
                            break

                    if not pchange_col_idx:
                        continue  # skip sheets without pChange

                    # Loop rows
                    for row_idx in range(2, sheet.max_row + 1):
                        cell = sheet.cell(row=row_idx, column=pchange_col_idx)
                        raw_val = cell.value
                        if raw_val is None:
                            continue

                        if isinstance(raw_val, str):
                            val_str = raw_val.replace("%", "").replace(",", "").strip()
                        else:
                            val_str = raw_val

                        try:
                            pChange_val = float(val_str)
                        except (TypeError, ValueError):
                            continue

                        fill_color = None
                        if pChange_val < 0:
                            fill_color = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        elif pChange_val > 0:
                            fill_color = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

                        if fill_color:
                            for col_idx in range(1, sheet.max_column + 1):
                                sheet.cell(row=row_idx, column=col_idx).fill = fill_color

            print(f"[SUCCESS] Exported all responses to {file_path}")
        except Exception as e:
            print(f"[ERROR] Failed to export Excel: {e}")

    def run(self):
        start_time = time.time()
        print(f"Execution started at: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(start_time))}")
        
        broad_indices_dict = {}
        gainers_dict = {}
        for market_index in self.marketIndices:
            broad_indices = self.fetch_broad_market_indices(market_index)
            broad_indices_dict[market_index] = broad_indices
            gainers_dict[market_index] = self.fetch_gainers_for_indices(market_index, broad_indices)
        self.export_to_excel(broad_indices_dict, gainers_dict)
        
        end_time = time.time()
        print(f"Execution ended at: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time))}")
        print(f"Total execution time: {end_time - start_time:.2f} seconds")


if __name__ == "__main__":
    exporter = NseTestDataExporter()
    exporter.run()
