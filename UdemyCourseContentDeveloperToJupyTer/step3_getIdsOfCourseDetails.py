import os
import json
import requests
import pandas as pd
from openpyxl import load_workbook


class UdemySupplementaryFetcher:
    def __init__(self, input_file, output_file, sheet_name, auth_file, base_folder):
        self.input_file = input_file
        self.output_file = output_file
        self.sheet_name = sheet_name
        self.auth_file = auth_file
        self.base_folder = base_folder

        # Ensure base folder exists
        os.makedirs(self.base_folder, exist_ok=True)

        # Load authentication data
        with open(self.auth_file, "r") as f:
            auth_data = json.load(f)
        self.ACCESS_TOKEN = auth_data.get("access_token")

        # Headers
        self.headers = {
            "accept": "application/json, text/plain, */*",
            "accept-language": "en-US",
            "x-requested-with": "XMLHttpRequest"
        }
        self.cookies = {"access_token": self.ACCESS_TOKEN}

    def read_course_ids(self):
        """Read course IDs from Excel (Column B starting Row 2)."""
        wb = load_workbook(self.input_file, data_only=True)
        ws = wb[self.sheet_name]

        course_ids = []
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            cell_value = row[0].value
            if cell_value is None:
                break
            course_ids.append(str(cell_value).strip())

        print("📚 Course IDs found:", course_ids)
        return course_ids

    def get_course_name(self, course_id):
        """Fetch course name for a given course_id."""
        url = f"https://www.udemy.com/api-2.0/courses/{course_id}/?fields[course]=id,title"
        response = requests.get(url, headers=self.headers, cookies=self.cookies)
        if response.status_code == 200:
            return response.json().get("title", "Unknown Course")
        return "Unknown Course"

    def get_curriculum_map(self, course_id):
        """Fetch curriculum to map lecture_id -> section_id, section_name, lecture_name."""
        url = f"https://www.udemy.com/api-2.0/courses/{course_id}/public-curriculum-items/?page_size=1000"
        response = requests.get(url, headers=self.headers, cookies=self.cookies)
        section_map = {}
        lecture_map = {}

        if response.status_code == 200:
            results = response.json().get("results", [])
            current_section_id, current_section_title = None, None

            for item in results:
                if item["_class"] == "chapter":
                    current_section_id = item.get("id")
                    current_section_title = item.get("title")
                    section_map[current_section_id] = current_section_title

                elif item["_class"] == "lecture":
                    lecture_map[item["id"]] = {
                        "lecture_title": item.get("title"),
                        "section_id": current_section_id,
                        "section_title": current_section_title,
                    }

        return lecture_map

    def fetch_assets(self, course_ids):
        """Fetch supplementary assets of type 'File' with full metadata."""
        all_data = []

        for course_id in course_ids:
            print(f"📥 Fetching course {course_id} ...")

            # Get course name
            course_name = self.get_course_name(course_id)

            # Get curriculum (map lecture -> section)
            lecture_map = self.get_curriculum_map(course_id)

            # Fetch lectures & supplementary files
            url = (
                f"https://www.udemy.com/api-2.0/users/me/subscribed-courses/{course_id}/lectures/"
                "?page_size=1000&fields[lecture]=id,title,asset,supplementary_assets"
            )

            response = requests.get(url, headers=self.headers, cookies=self.cookies)

            if response.status_code == 200:
                data = response.json()

                for lecture in data.get("results", []):
                    lecture_id = lecture.get("id")
                    lecture_info = lecture_map.get(lecture_id, {})

                    for asset in lecture.get("supplementary_assets", []):
                        if asset.get("asset_type") == "File":
                            all_data.append({
                                "course_id": course_id,
                                "course_name": course_name,
                                "section_id": lecture_info.get("section_id"),
                                "section_name": lecture_info.get("section_title"),
                                "lecture_id": lecture_id,
                                "lecture_name": lecture_info.get("lecture_title", lecture.get("title")),
                                "supplementary_asset_id": asset.get("id"),
                                "asset_title": asset.get("title"),
                            })
            else:
                print(f"❌ Failed for course {course_id}: {response.status_code}")

        return all_data

    def export_to_excel(self, all_data):
        """Export supplementary assets to Excel."""
        if all_data:
            df = pd.DataFrame(all_data)
            df.to_excel(self.output_file, index=False)
            print(f"✅ Exported to {self.output_file}")
        else:
            print("⚠️ No supplementary assets of type 'File' found.")

    def run(self):
        """Main execution flow."""
        course_ids = self.read_course_ids()
        all_data = self.fetch_assets(course_ids)
        self.export_to_excel(all_data)


# ======================= USAGE =======================
if __name__ == "__main__":
    base_folder = r"./udemyDownloads"
    input_file = os.path.join(base_folder, "udemy_courses.xlsx")
    output_file = os.path.join(base_folder, "udemy_CourseOutlineTitles.xlsx")
    auth_file = "Authentication.json"   # contains {"access_token": "XXXX"}
    sheet_name = "Udemy Courses"

    fetcher = UdemySupplementaryFetcher(input_file, output_file, sheet_name, auth_file, base_folder)
    fetcher.run()
